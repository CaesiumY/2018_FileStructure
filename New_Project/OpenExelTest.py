import openpyxl

wb = openpyxl.load_workbook("C:\\Users\\mn065\\Downloads\\NHIS_OPEN_GJ_2016.xlsx")
# wb = openpyxl.load_workbook("C:\\Users\\mn065\\Downloads\\new.xlsx")

ws = wb.active

print("load complete")

# ws_column = ws.max_column
# ws_row = ws.max_row
#
# print("column=")
# print(ws_column)
# print("row=")
# print(ws_row)

# for i in range(1, 4):
#     for j in range(1, 4):
#         b2 = ws.cell(row=i, column=j)
#         print(b2.value)

for i in range(1, 100):
    smoking_status = ws.cell(row=i, column=26)
    print(smoking_status.value)

wb.close()